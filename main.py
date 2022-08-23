#Importation des Bibliothèque
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#Creation d'un Fichier Excel avec les titres de chaque colonne
wb = Workbook()
ws = wb.active
ws.title = 'Posts'

heading = ['Title', 'Category', 'Date', 'Content', 'Link', 'Img']
ws.append(heading)

# Number of pages to scraping from Hespress
NUMBER_OF_PAGES = 6

print('[+] Start scraping Hespress')

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.60 Safari/537.36'
}

# 29423 last page of Hesspress
for i in range(1, NUMBER_OF_PAGES):
    print(f'Filtrer les données de chaque page {str(i)} :')

    #Requesting
    html_text = requests.get('https://www.hespress.com/?action=ajax_listing&paged=' + str(i) +
                             '&tq=MjAyMi0wNC0wNiAwMDowNTowMA%3D%3D&all_listing=1', headers=headers).text

    #Parsing
    soup = BeautifulSoup(html_text, 'lxml')
    posts = soup.find_all('div', class_='overlay card')

    #Scraping Content
    for post in posts:
        print(f'[+] Scraping post')
        post_title = post.find('h3', class_='card-title').text
        post_category = post.find('span', class_='cat').text
        post_img = post.find('img', class_='wp-post-image')['src']
        post_link = post.find('a', class_='stretched-link')['href']
        post_date = post.find('small', class_='text-muted time').text

    #Scraping Post Content    

    post_content_html_text = requests.get(post_link, headers=headers).text
    soup = BeautifulSoup(post_content_html_text, 'lxml')

    post_content_paragraphs = soup.find('div', class_='article-content').find_all('p')

    post_content = ''

    for p in post_content_paragraphs:
        post_content += p.text + '\n'

    # Add post to Excel file
    ws.append([post_title, post_category, post_date, post_content, post_link, post_img])

    # Add font to first line of Excel file
    for col in range(1, 7):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

    print('[+] Scraping post successfully')

# Save Excel file
wb.save('post.xlsx')
print('[+] Finish scraping Hesspress')